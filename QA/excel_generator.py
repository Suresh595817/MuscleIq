import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_enterprise_report(data, filename, suite_name):
    """
    Generates a highly detailed, multi-sheet Enterprise Excel report matching the requested format.
    data: list of dicts. Keys should match the columns: 
    ['Test ID', 'Module', 'Test Name', 'Priority', 'Preconditions', 'Steps', 'Test Data', 'Expected Result', 'Actual Result', 'Status', 'Duration (ms)', 'Device']
    """
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    pass_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
    fail_fill = PatternFill(start_color="EA4335", end_color="EA4335", fill_type="solid")
    pass_font = Font(bold=True, color="FFFFFF")
    fail_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Sheet 1: Executed Test Cases
    ws_all = wb.active
    ws_all.title = f"Executed Test Cases ({len(data)})"
    
    # Title Row
    ws_all.merge_cells('A1:L1')
    ws_all['A1'] = f"MuscleIQ - {suite_name} Test Execution Report ({len(data)} Cases)"
    ws_all['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_all['A1'].fill = PatternFill(start_color="202124", end_color="202124", fill_type="solid")
    ws_all['A1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ['Test ID', 'Module', 'Test Name', 'Priority', 'Preconditions', 'Steps', 'Test Data', 'Expected Result', 'Actual Result', 'Status', 'Duration (ms)', 'Device']
    
    # Header Row
    for col_num, header in enumerate(headers, 1):
        cell = ws_all.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border
        
        # Approximate column widths
        ws_all.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    ws_all.column_dimensions['F'].width = 40 # Steps
    ws_all.column_dimensions['H'].width = 40 # Expected
    ws_all.column_dimensions['I'].width = 40 # Actual

    passed_data = []
    
    # Data Rows
    for r_idx, row_data in enumerate(data, 3):
        for c_idx, key in enumerate(headers, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx)
            val = row_data.get(key, "")
            cell.value = val
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            
            if key == 'Status':
                if 'PASS' in str(val).upper():
                    cell.fill = pass_fill
                    cell.font = pass_font
                    if row_data not in passed_data:
                        passed_data.append(row_data)
                elif 'FAIL' in str(val).upper() or 'ERROR' in str(val).upper():
                    cell.fill = fail_fill
                    cell.font = fail_font

    # Sheet 2: Passed Tests
    ws_pass = wb.create_sheet(title=f"Passed Tests ({len(passed_data)})")
    # Copy Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_pass.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border
        ws_pass.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Data Rows for Passed
    for r_idx, row_data in enumerate(passed_data, 2):
        for c_idx, key in enumerate(headers, 1):
            cell = ws_pass.cell(row=r_idx, column=c_idx)
            val = row_data.get(key, "")
            cell.value = val
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if key == 'Status':
                cell.fill = pass_fill
                cell.font = pass_font

    # Sheet 3: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    metrics = [
        ("Total Tests Executed", len(data)),
        ("Passed Tests", len(passed_data)),
        ("Failed Tests", len(data) - len(passed_data)),
        ("Pass Rate (%)", f"{(len(passed_data)/len(data)*100 if len(data) > 0 else 0):.2f}%"),
        ("Environment", "GitHub Actions (Ubuntu/macOS)"),
        ("Target System", "MuscleIQ Live Web/App")
    ]
    for i, (k, v) in enumerate(metrics, 1):
        ws_metrics.cell(row=i, column=1).value = k
        ws_metrics.cell(row=i, column=1).font = Font(bold=True)
        ws_metrics.cell(row=i, column=2).value = v
        ws_metrics.column_dimensions['A'].width = 30
        ws_metrics.column_dimensions['B'].width = 40

    # Sheet 4: Module Summary
    ws_mod = wb.create_sheet(title="Module Summary")
    ws_mod.append(["Module", "Total Tests", "Passed", "Failed"])
    for cell in ws_mod[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    mod_stats = {}
    for d in data:
        mod = d.get('Module', 'Unknown')
        status = d.get('Status', '')
        if mod not in mod_stats:
            mod_stats[mod] = {'total': 0, 'pass': 0, 'fail': 0}
        mod_stats[mod]['total'] += 1
        if 'PASS' in str(status).upper():
            mod_stats[mod]['pass'] += 1
        else:
            mod_stats[mod]['fail'] += 1
            
    for row_idx, (mod, stats) in enumerate(mod_stats.items(), 2):
        ws_mod.cell(row=row_idx, column=1).value = mod
        ws_mod.cell(row=row_idx, column=2).value = stats['total']
        ws_mod.cell(row=row_idx, column=3).value = stats['pass']
        ws_mod.cell(row=row_idx, column=4).value = stats['fail']

    wb.save(filename)
